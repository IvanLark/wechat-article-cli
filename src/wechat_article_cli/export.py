"""运行结果导出

支持 JSON / CSV / Excel 三种格式。
默认导出到 wechat-article capability home 下的 `exports/runs/<run_id>/` 目录。
导出时从缓存读取文章内容。
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Literal

from wechat_article_cli.storage import Run

_HEADERS = ["文章标题", "公众号名称", "作者", "发布日期", "文章链接", "摘要", "文章内容"]


def _run_export_dir(run_id: str) -> Path:
    from wechat_article_cli._toolkit.runtime.home import get_data_dir

    d = get_data_dir("wechat_article") / "exports" / "runs" / run_id
    d.mkdir(parents=True, exist_ok=True)
    return d


def _get_content(link: str) -> str:
    """从缓存读取文章内容（markdown 格式）"""
    from wechat_article_cli.content import get_cached_content

    return get_cached_content(link, fmt="markdown") or ""


def _article_row(article: dict) -> list[str]:
    return [
        article.get("title", ""),
        article.get("account_name", ""),
        article.get("author", ""),
        article.get("create_date", ""),
        article.get("link", ""),
        article.get("digest", ""),
        _get_content(article.get("link", "")),
    ]


def export_run(
    run: Run, fmt: Literal["json", "csv", "excel"] = "json"
) -> Path:
    """导出 run 结果，返回输出文件路径"""
    out_dir = _run_export_dir(run.run_id)

    if fmt == "json":
        return _export_json(run, out_dir)
    elif fmt == "csv":
        return _export_csv(run, out_dir)
    elif fmt == "excel":
        return _export_excel(run, out_dir)
    else:
        raise ValueError(f"不支持的导出格式: {fmt}")


def _export_json(run: Run, out_dir: Path) -> Path:
    path = out_dir / "articles.json"
    data = {
        "run_id": run.run_id,
        "task_id": run.task_id,
        "created_at": run.created_at,
        "statistics": run.statistics.model_dump(),
        "articles": [
            {
                **article,
                "content": _get_content(article.get("link", "")),
            }
            for article in run.articles
        ],
    }
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def _export_csv(run: Run, out_dir: Path) -> Path:
    path = out_dir / "articles.csv"
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(_HEADERS)
        for article in run.articles:
            writer.writerow(_article_row(article))
    return path


def _export_excel(run: Run, out_dir: Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise RuntimeError("需要安装 openpyxl: uv add openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "文章数据"

    # 表头
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="366092")
    for col, header in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    even_fill = PatternFill("solid", fgColor="F5F5F5")
    for row_idx, article in enumerate(run.articles, 2):
        for col, val in enumerate(_article_row(article), 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = even_fill

    # 列宽
    widths = [40, 20, 12, 14, 60, 50, 80]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    ws.freeze_panes = "A2"

    path = out_dir / "articles.xlsx"
    wb.save(path)
    return path
